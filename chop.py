from openpyxl.reader.excel import load_workbook
import os
from openpyxl.worksheet.table import Table, TableStyleInfo
import win32com.client
import html
import os
from pprint import pprint

source = r"C:\Users\asorensen\OneDrive - CVRx Inc\2025_COMP_OPS\DocuSign_Process\STATUS REPORTS\Envelope Recipient Report 4.21.25.xlsx"
target_dir = r"C:\Users\asorensen\PycharmProjects\EXCEL_MANIPULATION\files"


def create_files(source_file, tab, column, target_directory):
    """Source file and tab correlate to the file you want to chop up. Column refers to the column that contains the
    groups that you want the individual files created for. Target directory is the location where you want
    to generate the chopped files. Note that you MUST delete all external connections (like queries) from
    the source before running, or the outputs will be corrupted."""
    # open the workbook
    workbook = load_workbook(source_file)
    sheet = workbook[tab]
    criteria = []
    col_index = None

    # add all distinct column values to the criteria variable and save the column letter to a variable
    # iterate over each column in the first row (header row)
    for col in sheet.iter_cols(min_row=1, max_row=1):
        # check if the column header matches the name provided as the column variable
        if col[0].value == column:
            col_index = col[0].col_idx
            # iterate over all cells in the column, skipping the header
            for cell in sheet[col[0].column_letter][1:]:
                if cell.value not in criteria:
                    criteria.append(cell.value)
                else:
                    continue
            break

    # iterate through each value in the criteria list and create a new workbook for each value
    for item in criteria:
        workbook = load_workbook(source_file)
        sheet = workbook[tab]

        # UPDATE THIS VARIABLE TO CHANGE THE NAME OF THE SAVED FILE
        save_path = os.path.join(target_directory, f"{item}.xlsx")

        # remove table formatting
        for table in list(sheet.tables):
            del sheet.tables[table]

        # delete rows that don't match the criteria
        for row in reversed(list(sheet.iter_rows(min_row=2, max_row=sheet.max_row))):
            if row[col_index - 1].value != item:
                sheet.delete_rows(row[0].row, 1)

        # format the remaining data as a new table
        max_row = sheet.max_row
        max_col = sheet.max_column
        table_range = f"A1:{chr(65 + max_col - 1)}{max_row}"
        table = Table(displayName=f"Table_1", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium16",
            showRowStripes=True
        )
        table.tableStyleInfo = style
        sheet.add_table(table)

        # save
        workbook.save(save_path)


def send_emails(files_directory, email_column, email_template):
    files_and_recipients = {}
    for filename in os.listdir(files_directory):
        filepath = os.path.join(files_directory, filename)
        workbook = load_workbook(filepath)
        sheet = workbook.active
        emails = []
        for col in sheet.iter_cols(min_row=1, max_row=1):
            if col[0].value == email_column:
                for cell in sheet[col[0].column_letter][1:]:
                    if cell.value not in emails:
                        emails.append(cell.value)
                    else:
                        continue
                break
        files_and_recipients[filepath] = emails
    # pprint(files_and_recipients)

    for key, value in files_and_recipients.items():
        outlook = win32com.client.Dispatch('Outlook.Application')
        mail = outlook.CreateItemFromTemplate(email_template)
        # mail = outlook.CreateItem(0)
        # mail.Body = """testing testing testing"""
        mail.Subject = f'Unsigned DocuSigns'
        mail.To = '; '.join(value)
        mail.CC = 'jmoore@cvrx.com'
        mail.Attachments.Add(fr'{key}')

        mail.Send()


create_files(source_file=source, tab="Sheet1", column='Region', target_directory=target_dir)
# send_emails(files_directory=target_dir, email_column="ASD Email", email_template=r"C:\Users\asorensen\PycharmProjects\EXCEL_MANIPULATION\eml.oft")